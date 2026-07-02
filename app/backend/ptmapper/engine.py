"""The deterministic PT mapping engine.

No AI. A brand file is read, its header located, columns mapped per profile,
values normalised through DB lookup tables, derived fields computed, and anything
unresolved is pushed to the review queue (never guessed). Re-running after a
lookup is added needs zero code change.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any

import openpyxl

from ptmapper.models import ControlledValue, ItemTaxonomy, Lookup, TaxonomyRule
from ptmapper.profiles import (
    ALIASES,
    CONTROLLED,
    FILENAME_HINTS,
    GENERIC_PROFILE,
    HEADER_KEYWORDS,
    PROFILES,
)

MAX_ROWS = 8000

# Provenance of a derived cell. 'direct' (normalised value IS a Master value) and
# 'derived' (deterministic, e.g. season-from-date) are trustworthy; the rest encode a
# judgement (an alias/seed mapping, a keyword rule, or a gender inferred from text) and
# are worth a steward's glance — that is what the UI should surface.
LOW_CONFIDENCE_SOURCES = {"alias", "rule", "inferred"}


@dataclass(frozen=True)
class BaseMappedFields:
    barcode: str
    design: str
    qty: float | None
    mrp: float | None
    desc: str
    raw_brand: str


@dataclass(frozen=True)
class PriceFields:
    prate: float | None
    basic: float | None
    input_tax: float | None


@dataclass(frozen=True)
class ResolvedFields:
    brand: str | None
    color: str | None
    size: str | None
    season: str | None
    taxonomy: dict


class UnsupportedFormat(Exception):
    pass


# ----------------------------------------------------------------------------- helpers
def norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip()).upper()


def raw_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip().rstrip("%")
    try:
        return float(s)
    except ValueError:
        return None


def money(v: Any) -> float | None:
    n = num(v)
    return round(n, 2) if n is not None else None


def clean_code(v: Any) -> str:
    """Stringify a barcode/HSN without a spurious trailing .0 from float cells."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%Y%m%d",
    "%m/%d/%Y",
]


def parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Excel serial date (common in .xlsb exports where dates arrive as numbers).
    if isinstance(v, (int, float)) and 20000 < float(v) < 80000:
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).date()
        except Exception:  # noqa: BLE001
            return None
    s = raw_str(v)
    if not s:
        return None
    s = s.split(".")[0] if s.isdigit() and len(s) > 8 else s
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def season_label(d: date) -> str:
    prefix = "SPRING SUMMER" if d.month <= 6 else "AUTUMN WINTER"
    return f"{prefix}({d.strftime('%b')}-{d.strftime('%y')})"


# --------------------------------------------------------------------------- normalisers
# Mechanical value normalisers: turn a brand's raw string into the KDPS canonical
# form *deterministically* so it lands directly on a Master-Sheet value. They do NOT
# encode business judgement (shade→bucket, category→item) — that lives in the DB
# lookup tables (grown by the review queue). Pure functions → unit-tested without a DB.

# Alpha size variants → KDPS master size. Master has XS/S/M/L/XL/XXL/3XL/4XL/5XL/6XL
# and "FREE SIZE" (no "2XL"/"XXXL"), so collapse those here.
_ALPHA_SIZE_ALIASES = {
    "2XL": "XXL",
    "XXXL": "3XL",
    "XXXXL": "4XL",
    "XXXXXL": "5XL",
    "XXXXXXL": "6XL",
    "FS": "FREE SIZE",
    "F": "FREE SIZE",
    "FREE": "FREE SIZE",
    "FREESIZE": "FREE SIZE",
    "ONESIZE": "FREE SIZE",
    "OS": "FREE SIZE",
    "1MTR": "FREE SIZE",
    "1MTR.": "FREE SIZE",
}
_AGE_Y_RE = re.compile(r"^(\d+)\s*-\s*(\d+)\s*(?:Y|YR|YRS|YEAR|YEARS)\.?$")
_AGE_M_RE = re.compile(r"^(\d+)\s*-\s*(\d+)\s*(?:M|MO|MTH|MTHS|MONTH|MONTHS)\.?$")
_PAREN_RE = re.compile(r"\(([^)]+)\)")
_ALPHA_SIZE_RE = re.compile(r"\d*X*[SML]")  # S, M, L, XS, XL, XXL, 3XL, ...


def _canon_alpha_size(s: str) -> str:
    u = re.sub(r"\s+", " ", s.strip().upper())
    return _ALPHA_SIZE_ALIASES.get(u.replace(" ", ""), u)


def _is_alpha_size(s: str) -> bool:
    return bool(_ALPHA_SIZE_RE.fullmatch(s.replace(" ", "")))


def _size_once(s: str) -> str:
    """One normalisation pass over an already upper/cleaned size string."""
    # "100 CMS" / "96 CMS." → drop the (possibly repeated) plural CMS suffix to the bare
    # number (a bare "96 CM" stays, since the Master keeps "96 CM" distinct from "96").
    s = re.sub(r"(?:\s*CMS\.?)+\s*$", "", s)
    if not s:
        return ""
    if s.replace(" ", "") in _ALPHA_SIZE_ALIASES:  # FS, 2XL, FREE, 1 MTR. ...
        return _ALPHA_SIZE_ALIASES[s.replace(" ", "")]
    m = _PAREN_RE.search(s)  # a measurement + an alpha size, either side in the parens
    if m:
        base = _canon_alpha_size(s[: m.start()])  # "XL (105 CMS)" → base "XL"
        inner = _canon_alpha_size(m.group(1))  # "1.14M(2XL)" → inner "XXL"
        if _is_alpha_size(base):
            return base
        if _is_alpha_size(inner):
            return inner
        return base or inner  # neither side alpha → reduced further on the next pass
    if "/" in s:  # "36/XS" = the SAME size in two notations → take the alpha (XS)
        parts = [p.strip() for p in s.split("/") if p.strip()]
        alpha = [p for p in parts if re.search(r"[A-Za-z]", p)]
        numeric = [p for p in parts if re.fullmatch(r"\d+", p)]
        if len(alpha) == 1 and numeric:  # num/alpha equivalence ("44/XXL" → "XXL")
            return _canon_alpha_size(alpha[0])
        return s  # ambiguous ("S/M" = two sizes, "36/38") → unchanged → review queue
    m = _AGE_Y_RE.match(s)  # "7-8Y" / "8-9 YEARS" → "7-8 Y"
    if m:
        return f"{int(m.group(1))}-{int(m.group(2))} Y"
    m = _AGE_M_RE.match(s)  # "0-6M" / "12-18 MONTHS" → "0-6 M"
    if m:
        return f"{int(m.group(1))}-{int(m.group(2))} M"
    return _canon_alpha_size(s)


def normalize_size(raw: Any) -> str:
    """Brand size string → KDPS master size (best effort; unresolved stays for review).

    Handles float artifacts ('44.0'→'44'), free-size words, metric-with-alpha
    ('1.14M(2XL)'→'XXL', '96CM(M)'→'M'), slash duals ('36/XS'→'XS'), and age
    bands ('7-8Y'/'8-9 YEARS'→'7-8 Y'). Plain numbers/bra sizes pass through.

    Idempotent: a paren/slash can expose a further-reducible token (e.g. '(105 CMS)'),
    so passes are repeated to a fixed point — f(f(x)) == f(x), so re-running the engine
    never drifts a value.
    """
    s = clean_code(raw).upper().strip()
    for _ in range(4):
        nxt = _size_once(s)
        if nxt == s:
            return nxt
        s = nxt
    return s


_COLOR_LEAD_CODE = re.compile(r"^\s*\d+\s*[-_]\s*")  # "88-BEIGE", "16 - BLUE"
_COLOR_TRAIL_NUM = re.compile(r"\s*\d+\s*(?:/\s*\d+)?\s*$")  # "BLACK73", "WHITE74/1"
_COLOR_TRAIL_TOK = re.compile(r"\s+(?:DN|DBY|DB|MEL|MELANGE)\s*$")  # denim/melange wash codes


def normalize_color(raw: Any) -> str:
    """Strip a brand colour string down to its shade word (mechanical only):
    leading numeric codes ('88-BEIGE'→'BEIGE'), trailing wash/lot codes
    ('BLACK73/1'→'BLACK', 'BLUE DN89'→'BLUE'). The shade→23-bucket judgement
    ('MID BLUE'→'BLUE') lives in the colour lookup table, not here.
    """
    s0 = re.sub(r"\s+", " ", raw_str(raw).upper()).strip()
    if not s0:
        return ""
    # strip leading codes + trailing lot numbers / wash codes to a fixed point, so the
    # result is idempotent (e.g. "L BLU DN88" → "L BLU", "BLACK73/1" → "BLACK").
    s = s0
    prev = None
    while prev != s:
        prev = s
        s = _COLOR_LEAD_CODE.sub("", s)
        s = _COLOR_TRAIL_NUM.sub("", s)
        s = _COLOR_TRAIL_TOK.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    # An all-code input ("501", "123-") must NOT vanish to blank — keep it reviewable
    # (a blank colour records no miss, so the value would be silently dropped).
    return s or s0


_GENDER_KEYWORDS = (
    # (keyword, KDPS gender) — order matters: kids/female checked before adult/male
    (" GIRLS ", "KIDS FEMALE"),
    (" GIRL ", "KIDS FEMALE"),
    (" BOYS ", "KIDS MALE"),
    (" BOY ", "KIDS MALE"),
    (" INFANT ", "UNISEX"),
    (" WOMENS ", "FEMALE"),
    (" WOMEN ", "FEMALE"),
    (" WOMAN ", "FEMALE"),
    (" LADIES ", "FEMALE"),
    (" LADIE ", "FEMALE"),
    (" FEMALE ", "FEMALE"),
    (" MENS ", "MALE"),
    (" MEN ", "MALE"),
    (" MAN ", "MALE"),
    (" GENTS ", "MALE"),
    (" MALE ", "MALE"),
    (" KIDS ", "UNISEX"),
    (" KID ", "UNISEX"),
    (" JUNIOR ", "UNISEX"),
    (" UNISEX ", "UNISEX"),
)


def gender_from_text(text: str) -> str:
    """Infer KDPS gender from words in a description (fallback to a gender column).
    Separators are flattened so 'SOCKS-MENS' and 'Senior Girls Top' both match.
    """
    flat = " " + re.sub(r"[^A-Z0-9]+", " ", norm(text)).strip() + " "
    for kw, g in _GENDER_KEYWORDS:
        if kw in flat:
            return g
    return ""


_FIT_CODE_TOKEN = re.compile(r"^[A-Z]{2,3}$")


def fit_code_candidates(raw: Any) -> list[str]:
    """Candidates for a coded FIT TYPE column (profile flag ``fit_code_tokens``).

    Peter England / ABFRL encode fit as '<line> <family> <style…>' — e.g.
    'PJ RG OCTANEMIDSTR', 'PC SL Snug', 'PC RG Regular': token 2 is the fit
    family (RG = Regular, SL = Slim). Order: the full value first (a plain
    'SLIM FIT' still hits the normal lookup), then token 2, then a trailing
    word ('Regular'/'Slim'). Unknown codes resolve to nothing → review queue.
    """
    s = norm(raw)
    if not s:
        return []
    candidates = [s]
    tokens = s.split()
    # The coded form always has ≥3 tokens — a plain 2-word fit ("SLIM FIT")
    # must stay a single candidate for the normal lookup.
    if len(tokens) >= 3 and _FIT_CODE_TOKEN.fullmatch(tokens[1]):
        candidates.append(tokens[1])
    if len(tokens) >= 3 and tokens[-1].isalpha():
        candidates.append(tokens[-1])
    return candidates


# ----------------------------------------------------------------------------- reading
def _read_xlsx(content: bytes) -> list[tuple[str, list[list]]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[tuple[str, list[list]]] = []
    for ws in wb.worksheets:
        rows: list[list] = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(r))
            if i >= MAX_ROWS:
                break
        out.append((ws.title, rows))
    wb.close()
    return out


def _read_csv(content: bytes) -> list[tuple[str, list[list]]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UnsupportedFormat("Could not decode the CSV text.")
    rows = list(csv.reader(io.StringIO(text)))[: MAX_ROWS + 1]
    return [("csv", rows)]


def _read_xls(content: bytes) -> list[tuple[str, list[list]]]:
    """Legacy OLE2 .xls via xlrd; Excel serial dates → datetime."""
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    out: list[tuple[str, list[list]]] = []
    for sh in book.sheets():
        rows: list[list] = []
        for i in range(min(sh.nrows, MAX_ROWS + 1)):
            row: list = []
            for j in range(sh.ncols):
                cell = sh.cell(i, j)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        v = xlrd.xldate_as_datetime(cell.value, book.datemode)
                    except Exception:  # noqa: BLE001
                        pass
                row.append(v)
            rows.append(row)
        out.append((sh.name, rows))
    return out


def _read_xlsb(content: bytes) -> list[tuple[str, list[list]]]:
    """Binary .xlsb (e.g. Madura SAP export) via pyxlsb."""
    from pyxlsb import open_workbook as open_xlsb

    out: list[tuple[str, list[list]]] = []
    with open_xlsb(io.BytesIO(content)) as wb:
        for name in wb.sheets:
            rows: list[list] = []
            with wb.get_sheet(name) as sheet:
                for i, row in enumerate(sheet.rows()):
                    rows.append([c.v for c in row])
                    if i >= MAX_ROWS:
                        break
            out.append((name, rows))
    return out


def read_sheets(content: bytes, filename: str, content_type: str) -> list[tuple[str, list[list]]]:
    name = (filename or "").lower()
    is_zip = content[:4] == b"PK\x03\x04"  # xlsx / xlsb are both ZIP containers
    is_ole = content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy OLE2 .xls

    # .xlsb is a ZIP too, so it must be matched by extension before the ZIP sniff.
    if name.endswith(".xlsb"):
        return _read_xlsb(content)
    # Real container wins over a (possibly wrong) extension — covers .xls/.csv that
    # are actually .xlsx, and .csv/.xls that are actually OLE2 .xls.
    if is_zip:
        return _read_xlsx(content)
    if is_ole:
        return _read_xls(content)
    if name.endswith(".xlsx") or "spreadsheetml" in (content_type or ""):
        return _read_xlsx(content)
    if name.endswith(".xls"):
        return _read_xls(content)
    if name.endswith(".csv") or "csv" in (content_type or ""):
        return _read_csv(content)
    raise UnsupportedFormat(
        "Unsupported file type. Upload a .xlsx, .xls, .xlsb or .csv brand PT file."
    )


# ----------------------------------------------------------------------------- detection
def score_header(row: list) -> float:
    s = 0.0
    for c in row:
        cv = raw_str(c).lower()
        if not cv:
            continue
        if cv in HEADER_KEYWORDS:
            s += 1
        elif len(cv) < 40 and any(k in cv for k in HEADER_KEYWORDS):
            s += 0.5
    return s


def detect_header(rows: list[list]) -> int:
    best_i, best = 0, -1.0
    for i in range(min(25, len(rows))):  # printed-invoice item tables can start deep
        sc = score_header(rows[i])
        if sc > best:
            best, best_i = sc, i
    return best_i


def _has_data(rows: list[list]) -> bool:
    return any(any(raw_str(c) for c in r) for r in rows[:30])


def choose_sheet(sheets: list[tuple[str, list[list]]]) -> tuple[str, list[list]]:
    # Prefer a sheet whose name matches a profile's sheet_contains token.
    for p in PROFILES:
        sc = p["match"].get("sheet_contains")
        if sc:
            for name, rows in sheets:
                if sc in name.upper() and _has_data(rows):
                    return name, rows
    # Else the non-empty sheet with the best-scoring header row.
    best = None
    best_score = -1.0
    for name, rows in sheets:
        if not _has_data(rows):
            continue
        hi = detect_header(rows)
        sc = score_header(rows[hi]) + min(len(rows), 100) * 0.001
        if sc > best_score:
            best_score, best = sc, (name, rows)
    return best or sheets[0]


def profile_by_code(code: str) -> dict:
    for p in PROFILES:
        if p["code"] == code:
            return p
    return GENERIC_PROFILE


def identify_profile(filename: str, header_set: set[str], sheet_name: str) -> dict:
    up = (filename or "").upper()
    for kw, code in FILENAME_HINTS.items():
        if kw in up:
            p = profile_by_code(code)
            need = p.get("match", {}).get("header_has")
            # Use the filename hint only if the file's columns actually fit that profile;
            # otherwise a simple sheet named like a wide-format brand would be forced into
            # the wrong profile and drop every row. Fall through to fingerprinting instead.
            if not need or all(h in header_set for h in need):
                return p
    for p in PROFILES:
        m = p["match"]
        if not m:
            continue
        if m.get("sheet_contains") and m["sheet_contains"] not in (sheet_name or "").upper():
            continue
        if m.get("header_has") and not all(h in header_set for h in m["header_has"]):
            continue
        return p
    return GENERIC_PROFILE


def build_records(rows: list[list], header_idx: int) -> list[dict[str, Any]]:
    headers = [norm(c) for c in rows[header_idx]]
    records: list[dict[str, Any]] = []
    blanks_in_a_row = 0
    for r in rows[header_idx + 1 :]:
        if not any(raw_str(c) for c in r):
            blanks_in_a_row += 1
            if blanks_in_a_row >= 3:
                break
            continue
        blanks_in_a_row = 0
        first = raw_str(r[0]).lower() if r else ""
        if first.startswith("total") or first.startswith("grand total"):
            break
        rec: dict[str, Any] = {}
        for j, h in enumerate(headers):
            if h and j < len(r):
                rec.setdefault(h, r[j])
        records.append(rec)
    return records


# ----------------------------------------------------------------------------- resolver
class Resolver:
    """DB-backed value normaliser; records every miss for the review queue."""

    def __init__(self) -> None:
        self.controlled: dict[str, set[str]] = {}
        # normalised key → exact master value, so a normaliser whose output already
        # IS a master value resolves directly (no lookup row needed).
        self.controlled_exact: dict[str, dict[str, str]] = {}
        for cv in ControlledValue.objects.all():
            self.controlled.setdefault(cv.dimension, set()).add(cv.value)
            self.controlled_exact.setdefault(cv.dimension, {})[norm(cv.value)] = cv.value
        # keyed by (dimension, brand): a brand-scoped rule ("" = global) so the
        # resolver can prefer a Peter England / Zilu rule over the global fallback.
        self.lookups: dict[tuple[str, str], dict[str, str]] = {}
        for lk in Lookup.objects.all():
            self.lookups.setdefault((lk.dimension, lk.brand), {})[lk.source_key] = lk.target_value
        self.item_taxonomy = {
            it.item: (it.sub_category, it.type) for it in ItemTaxonomy.objects.all()
        }
        self.rules = list(TaxonomyRule.objects.all())  # ordered -priority
        self.misses: dict[tuple[str, str], dict] = {}

    def _miss(self, dimension: str, raw: str, ctx: dict | None = None) -> None:
        raw = raw.strip()[:300]
        if not raw:
            return
        rec = self.misses.setdefault(
            dimension + "|" + raw,
            {"dimension": dimension, "raw": raw, "occurrences": 0, "samples": [], "brands": []},
        )
        rec["occurrences"] += 1
        if ctx:
            if len(rec["samples"]) < 5:
                sample = ctx.get("desc") or ctx.get("brand") or ""
                if sample and sample not in rec["samples"]:
                    rec["samples"].append(sample[:120])
            # the row's *resolved* brand(s): lets a review resolution default to
            # brand scope when a miss came from exactly one brand (D4).
            sb = ctx.get("scope_brand") or ""
            if sb and sb not in rec["brands"]:
                rec["brands"].append(sb)

    def _resolve(
        self, dimension: str, candidate: Any, ctx: dict | None = None, brand: str = ""
    ) -> tuple:
        """Layered resolve → ``(value, source)``. Order: normalise → direct Master hit
        ('direct', high confidence) → brand-scoped alias → global alias ('alias', a
        review-able judgement) → '' (a miss, recorded for the queue). A brand-scoped
        rule wins over a global one so a per-brand fix never leaks to other brands."""
        key = norm(candidate)
        if not key:
            return None, ""
        exact = self.controlled_exact.get(dimension, {}).get(key)
        if exact:
            return exact, "direct"
        if brand:
            hit = self.lookups.get((dimension, brand), {}).get(key)
            if hit:
                return hit, "alias"
        hit = self.lookups.get((dimension, ""), {}).get(key)
        if hit:
            return hit, "alias"
        self._miss(dimension, str(candidate), ctx)
        return None, ""

    def brand(self, raw_value: str, ctx=None) -> tuple:
        return self._resolve("brand", raw_value, ctx)  # the brand dimension is global-only

    def color(self, raw_value: str, ctx=None, brand: str = "") -> tuple:
        return self._resolve("color", normalize_color(raw_value), ctx, brand)

    def size(self, raw_value: Any, ctx=None, brand: str = "") -> tuple:
        return self._resolve("size", normalize_size(raw_value), ctx, brand)

    def fit(
        self,
        raw_value: str,
        ctx=None,
        candidates: list[str] | None = None,
        brand: str = "",
    ) -> tuple:
        if not candidates:
            return self._resolve("fit", raw_value, ctx, brand)
        # Coded fit column: try each candidate silently (brand-scoped alias before
        # global); one miss (the original raw) only when every candidate fails — a
        # partial code must not clutter the queue.
        for i, cand in enumerate(candidates):
            key = norm(cand)
            if not key:
                continue
            exact = self.controlled_exact.get("fit", {}).get(key)
            if exact:
                return exact, ("direct" if i == 0 else "alias")
            if brand:
                hit = self.lookups.get(("fit", brand), {}).get(key)
                if hit:
                    return hit, "alias"
            hit = self.lookups.get(("fit", ""), {}).get(key)
            if hit:
                return hit, "alias"
        self._miss("fit", raw_str(raw_value), ctx)
        return None, ""

    def brand_gender(self, brand: str) -> str:
        """Brand-level default gender (seeded for unambiguous single-gender brands);
        the LAST gender fallback — see _map_taxonomy. Keyed by the brand as its
        source_key, so it stays a global ("") lookup row."""
        return self.lookups.get(("brand_gender", ""), {}).get(norm(brand), "")

    def gender(self, raw_value: str) -> tuple:
        key = norm(raw_value)
        if not key:
            return "", ""
        exact = self.controlled_exact.get("gender", {}).get(key)
        if exact:  # a real Master gender ("MALE") is high-confidence, not an alias
            return exact, "direct"
        hit = self.lookups.get(("gender", ""), {}).get(key)  # gender aliases are global
        if hit:  # an alias ("MEN"→MALE, "GIRLS"→KIDS FEMALE)
            return hit, "alias"
        return "", ""

    def season_from_date(self, d: date):
        label = season_label(d)
        if label in self.controlled.get("season", set()):
            return label
        return None

    def season_from_code(self, code: str, brand: str = ""):
        """Resolve an explicit brand season code via the lookup table only (brand-scoped
        alias before global). Returns None (without recording a miss) when unknown — the
        caller decides whether to fall back to the invoice date before logging a miss."""
        key = norm(code)
        if not key:
            return None
        if brand:
            hit = self.lookups.get(("season", brand), {}).get(key)
            if hit:
                return hit
        return self.lookups.get(("season", ""), {}).get(key)

    def taxonomy(self, desc: str, ctx=None) -> dict:
        # Whole-word/phrase match for every pattern ("PANT" hits "TRACK PANT" but never
        # "PANTIE"/"LAPTOP"). A glued substring match is allowed ONLY inside an SAP
        # "finished-goods" token (Madura packs the item into 'FGTROUSER'/'FGKJEANS'); this
        # recovers those without the plain-English false positives a global substring match
        # caused (e.g. 'FLOWER' contains 'LOWER', 'ADDRESS' contains 'DRESS').
        spaced = re.sub(r"[^A-Z0-9]+", " ", norm(desc)).strip()
        flat = " " + spaced + " "
        fg_tokens = [t for t in spaced.split() if t.startswith("FG")]
        for rule in self.rules:
            p = re.sub(r"[^A-Z0-9]+", " ", norm(rule.pattern)).strip()
            if not p:
                continue
            glued_ok = " " not in p and len(p) >= 5 and any(p in t for t in fg_tokens)
            if f" {p} " in flat or glued_ok:
                return {
                    "gender": rule.gender,
                    "sub_category": rule.sub_category,
                    "type": rule.type,
                    "item": rule.item,
                    "fit": rule.fit,
                }
        if desc.strip():
            self._miss("taxonomy", desc, ctx or {"desc": desc})
        return {}


# ----------------------------------------------------------------------------- mapping
def _pick(rec: dict, role: str, profile: dict) -> Any:
    override = profile.get("overrides", {}).get(role)
    if override:
        if isinstance(override, list):
            return [rec.get(norm(c)) for c in override]
        return rec.get(norm(override))
    for alias in ALIASES.get(role, []):
        if alias in rec and raw_str(rec[alias]):
            return rec[alias]
    return None


def _build_desc(rec: dict, profile: dict) -> str:
    cols = profile.get("overrides", {}).get("DESC_SRC") or ALIASES["DESC_SRC"]
    parts = []
    for c in cols:
        v = raw_str(rec.get(norm(c)))
        if v:
            parts.append(v)
    return " ".join(parts)


def _getter(rec: dict, profile: dict) -> Callable[[str], Any]:
    return lambda role: _pick(rec, role, profile)


def _extract_base_fields(
    rec: dict,
    profile: dict,
    g: Callable[[str], Any],
    brand_default: str,
) -> BaseMappedFields:
    return BaseMappedFields(
        barcode=clean_code(g("BARCODE")),
        design=raw_str(g("DESIGN")),
        qty=num(g("QTY")),
        mrp=money(g("MRP")),
        desc=_build_desc(rec, profile),
        raw_brand=raw_str(g("BRAND_SRC")) or profile.get("brand_const", "") or brand_default,
    )


def _has_mappable_payload(base: BaseMappedFields) -> bool:
    return bool((base.barcode or base.design) and (base.qty is not None or base.mrp is not None))


def _map_prices(rec: dict, profile: dict, g, qty: float | None) -> tuple:
    """(P RATE, BASIC, INPUT TAX). BASIC may be derived per-unit from taxable amount."""
    prate = money(g("PRATE_SRC"))
    if profile.get("flags", {}).get("basic_from_taxable_per_unit"):
        tax_amt = num(rec.get("TAXABLE_AMOUNT"))
        basic = round(tax_amt / qty, 2) if (tax_amt and qty) else money(g("BASIC_SRC"))
    else:
        basic = money(g("BASIC_SRC"))
    return prate, basic, num(g("TAX_SRC"))


def _price_fields(rec: dict, profile: dict, g, qty: float | None) -> PriceFields:
    prate, basic, input_tax = _map_prices(rec, profile, g, qty)
    return PriceFields(prate=prate, basic=basic, input_tax=input_tax)


def _map_season(
    resolver: Resolver, g, ctx: dict, ctx_date: date | None = None, brand: str = ""
) -> tuple:
    """SEASON → ``(value, source)``. The explicit season code wins ('alias') — a season
    is a NAME, never a date; the invoice date is the fallback ('derived'), then the
    operator-supplied invoice date from upload context ('derived'). A single miss is
    recorded only when *everything* fails (an unknown code alone shouldn't clutter the
    queue when a date resolves it)."""
    code = raw_str(g("SEASON_SRC"))
    season = resolver.season_from_code(code, brand)
    if season:
        return season, "alias"
    d = parse_date(g("DATE_SRC"))
    if d:
        season = resolver.season_from_date(d)
        if season:
            return season, "derived"
    if ctx_date:
        season = resolver.season_from_date(ctx_date)
        if season:
            return season, "derived"
    subject = code or (d.isoformat() if d else "")
    if subject:
        resolver._miss("season", subject, ctx)
    return None, ""


def _map_taxonomy(
    resolver: Resolver,
    desc: str,
    g,
    ctx: dict,
    prov: dict,
    profile: dict | None = None,
    brand: str | None = None,
) -> dict:
    """The 5-axis merchandising grid (+ ITEM-suggested sub/type) for one row, recording
    each axis's provenance in ``prov``.

    GENDER: a gender column wins ('direct'/'alias'), then a gender word in the
    description ('inferred'), then a rule's default ('rule'), then the resolved
    brand's seeded default gender ('inferred'). FIT: an explicit fit column
    ('direct'/'alias') beats a rule's fit ('rule'). ITEM / SUB CATEGORY / TYPE
    come from the matched rule + ITEM→helper ('rule').
    """
    tax = resolver.taxonomy(desc, ctx)
    gender, g_src = resolver.gender(raw_str(g("GENDER_SRC")))
    if not gender:
        from_text = gender_from_text(desc)
        if from_text:
            gender, g_src = from_text, "inferred"
        elif tax.get("gender"):
            gender, g_src = tax["gender"], "rule"
        elif brand:
            from_brand = resolver.brand_gender(brand)
            if from_brand:
                gender, g_src = from_brand, "inferred"
    prov["GENDER"] = g_src if gender else ""

    item = tax.get("item", "")
    prov["ITEM"] = "rule" if item else ""

    fit_raw = raw_str(g("FIT_SRC"))
    candidates = None
    if profile and profile.get("flags", {}).get("fit_code_tokens"):
        candidates = fit_code_candidates(fit_raw)
    fit, fit_src = resolver.fit(fit_raw, ctx, candidates=candidates, brand=brand or "")
    if not fit and tax.get("fit"):
        fit, fit_src = tax["fit"], "rule"
    prov["FIT"] = fit_src if fit else ""

    sub, typ = tax.get("sub_category", ""), tax.get("type", "")
    sub_src = "rule" if sub else ""
    typ_src = "rule" if typ else ""
    sug_sub = sug_type = ""
    if item and item in resolver.item_taxonomy:
        sug_sub, sug_type = resolver.item_taxonomy[item]
        if not sub:
            sub, sub_src = sug_sub, "rule"
        if not typ:
            typ, typ_src = sug_type, "rule"
    prov["SUB CATEGORY"] = sub_src if sub else ""
    prov["TYPE"] = typ_src if typ else ""
    return {
        "gender": gender,
        "item": item,
        "fit": fit,
        "sub": sub,
        "typ": typ,
        "sug_sub": sug_sub,
        "sug_type": sug_type,
    }


_COLOR_CODE_PREFIX = re.compile(r"^\s*\d+\s*-\s*")


def _color_src(g: Callable[[str], Any], profile: dict) -> str:
    """COLOR source value, with a leading numeric code prefix stripped for archetypes
    that encode colour as '<code>-<NAME>' (Ginesys CATEGORY3, e.g. '16-BLUE' → 'BLUE').
    Plain colour names and unflagged profiles pass through unchanged; an unresolved
    colour still goes to the review queue (never guessed)."""
    raw = raw_str(g("COLOR_SRC"))
    if profile.get("flags", {}).get("color_strip_code_prefix"):
        return _COLOR_CODE_PREFIX.sub("", raw)
    return raw


def _resolve_fields(
    resolver: Resolver,
    g: Callable[[str], Any],
    base: BaseMappedFields,
    profile: dict,
    prov: dict,
    ctx_brand: str = "",
    ctx_date: date | None = None,
) -> ResolvedFields:
    ctx = {"brand": base.raw_brand, "desc": base.desc}
    brand_v, brand_s = resolver.brand(base.raw_brand, ctx)
    if not brand_v and ctx_brand:
        # Operator context at upload — the structural fix for brand-less files
        # (Madura/Jockey/printed invoices). Deterministic, so 'derived'.
        brand_v, brand_s = ctx_brand, "derived"
    # Scope for the other dimensions = this row's resolved BRAND (D4). Multi-brand
    # files (Madura AS-VH-LP, Arvind) therefore scope each row independently. Threaded
    # into the miss ctx so a review resolution can default to the right brand.
    scope_brand = brand_v or ctx_brand or ""
    sctx = {**ctx, "scope_brand": scope_brand}
    color_v, color_s = resolver.color(_color_src(g, profile), sctx, brand=scope_brand)
    size_v, size_s = resolver.size(g("SIZE_SRC"), sctx, brand=scope_brand)
    season_v, season_s = _map_season(resolver, g, sctx, ctx_date, brand=scope_brand)
    prov["BRAND"] = brand_s if brand_v else ""
    prov["COLOR"] = color_s if color_v else ""
    prov["SIZE"] = size_s if size_v else ""
    prov["SEASON"] = season_s if season_v else ""
    return ResolvedFields(
        brand=brand_v,
        color=color_v,
        size=size_v,
        season=season_v,
        taxonomy=_map_taxonomy(resolver, base.desc, g, sctx, prov, profile, scope_brand),
    )


def _margin(mrp: float | None, prate: float | None) -> float | str:
    return round((mrp - prate) / mrp * 100, 2) if (mrp and prate and mrp > 0) else ""


def _quantity_out(qty: float | None) -> int | str:
    return int(qty) if qty else ""


def _build_kdps_row(
    base: BaseMappedFields,
    prices: PriceFields,
    resolved: ResolvedFields,
    g: Callable[[str], Any],
) -> dict:
    tx = resolved.taxonomy
    input_tax = prices.input_tax if prices.input_tax is not None else ""
    return {
        "SEASON": resolved.season or "",
        "BRAND": resolved.brand or "",
        "COLOR": resolved.color or "",
        "GENDER": tx["gender"] or "",
        "SUB CATEGORY": tx["sub"] or "",
        "TYPE": tx["typ"] or "",
        "ITEM": tx["item"] or "",
        "FIT": tx["fit"] or "",
        "SIZE": resolved.size or "",
        "BARCODE": base.barcode,
        "DESIGN": base.design,
        "HSN": clean_code(g("HSN")),
        "QTY": _quantity_out(base.qty),
        "MRP": base.mrp if base.mrp is not None else "",
        "BASIC": prices.basic if prices.basic is not None else "",
        "P RATE": prices.prate if prices.prate is not None else "",
        "INPUT TAX": input_tax,
        "OUTPUT TAX": input_tax,
        "NAG": _quantity_out(base.qty),
        "MARGIN": _margin(base.mrp, prices.prate),
        "SUGGESTED SUB CATEGORY": tx["sug_sub"],
        "SUGGESTED TYPE": tx["sug_type"],
    }


def _blank_controlled_columns(row: dict) -> list:
    return [column for column in CONTROLLED if not row[column]]


def _raw_sources(base: BaseMappedFields, g: Callable[[str], Any], profile: dict) -> dict:
    """The raw source value behind each controlled column (what map_record read),
    stored per row so an editor can bulk-apply a fix to 'every row with this raw
    value' and a future 'learn this edit' can seed a Lookup from it."""
    desc = base.desc[:300]
    fit_raw = raw_str(g("FIT_SRC"))
    return {
        "SEASON": (raw_str(g("SEASON_SRC")) or raw_str(g("DATE_SRC")))[:300],
        "BRAND": base.raw_brand[:300],
        "COLOR": _color_src(g, profile)[:300],
        "GENDER": (raw_str(g("GENDER_SRC")) or desc)[:300],
        "SUB CATEGORY": desc,
        "TYPE": desc,
        "ITEM": desc,
        "FIT": (fit_raw or desc)[:300],
        "SIZE": raw_str(g("SIZE_SRC"))[:300],
    }


def map_record(
    rec: dict,
    profile: dict,
    resolver: Resolver,
    brand_default: str,
    ctx_brand: str = "",
    ctx_date: date | None = None,
) -> tuple[dict, list, dict, dict] | None:
    g = _getter(rec, profile)
    base = _extract_base_fields(rec, profile, g, brand_default)
    if not _has_mappable_payload(base):
        return None
    prov: dict[str, str] = {}
    prices = _price_fields(rec, profile, g, base.qty)
    resolved = _resolve_fields(resolver, g, base, profile, prov, ctx_brand, ctx_date)
    row = _build_kdps_row(base, prices, resolved, g)
    return row, _blank_controlled_columns(row), prov, _raw_sources(base, g, profile)


def run_mapping(
    content: bytes, filename: str, content_type: str, context: dict | None = None
) -> dict:
    """Map a brand file into KDPS rows + review misses. Pure read (no DB writes).

    ``context`` is optional operator input from upload time — ``{"brand": <Master
    brand>, "invoice_date": <ISO date>}`` — used only as fallbacks when the file
    itself yields no brand / no date (provenance 'derived')."""
    sheets = read_sheets(content, filename, content_type)
    sheet_name, rows = choose_sheet(sheets)
    if not any(any(raw_str(c) for c in r) for r in rows):
        raise UnsupportedFormat("This file has no data rows to map.")
    truncated = len(rows) > MAX_ROWS
    header_idx = detect_header(rows)
    headers = [norm(c) for c in rows[header_idx]]
    header_set = {h for h in headers if h}
    profile = identify_profile(filename, header_set, sheet_name)

    ctx_brand = raw_str((context or {}).get("brand"))
    ctx_date = parse_date((context or {}).get("invoice_date"))

    # brand fallback from the filename (alpha words only) for files lacking a brand
    # column — skipped when the operator supplied the brand (context wins over a
    # filename guess, and the guess's inevitable miss would only clutter the queue).
    stem = (filename or "").rsplit(".", 1)[0]
    brand_default = (
        "" if ctx_brand else norm(" ".join(t for t in re.split(r"[ _\-]+", stem) if t.isalpha()))
    )

    resolver = Resolver()
    records = build_records(rows, header_idx)
    kdps_rows: list[dict] = []
    line_no = 0
    blank_cells = 0
    low_confidence_cells = 0
    for rec in records:
        mapped = map_record(rec, profile, resolver, brand_default, ctx_brand, ctx_date)
        if mapped is None:
            continue
        line_no += 1
        row, blanks, prov, raw = mapped
        blank_cells += len(blanks)
        low_confidence_cells += sum(1 for s in prov.values() if s in LOW_CONFIDENCE_SOURCES)
        kdps_rows.append(
            {"line_no": line_no, "data": row, "blanks": blanks, "provenance": prov, "raw": raw}
        )

    reviews = list(resolver.misses.values())
    return {
        "profile_code": profile["code"],
        "profile_name": profile["name"],
        "archetype": profile["archetype"],
        "brand_guess": brand_default,
        "meta": {
            "sheet": sheet_name,
            "header_row": header_idx,
            "headers": [h for h in headers if h],
            "source_rows": len(records),
            "truncated": truncated,
            "row_limit": MAX_ROWS,
        },
        "rows": kdps_rows,
        "reviews": reviews,
        "blank_cells": blank_cells,
        "low_confidence_cells": low_confidence_cells,
    }
